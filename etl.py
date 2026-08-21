"""
Steinhoff International — Question 4
Extract, transform and validate layer for the Version 2 capture workbook.

Version 2 restates the whole series in rand, extends cash flow to FY2017, and adds
an Altman Z'' panel, partial Beneish indices and share price anchor points. This
module reads those sheets and produces:

    statements   tidy long frame of every captured statement line
    ratios       recomputed from the statement lines, not read from the Ratios sheet
    checks       integrity findings, disclosed rather than corrected
    events       the red-flag chronology
    scorecard    traffic-light status per indicator per year
    prices       share price anchor points
    distress     Altman Z'' scores and zones as captured
    beneish      partial manipulation indices as captured
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pandas as pd

# Version 2 restates every year in rand, so the currency break is gone.
CURRENCY = "ZAR"
PHASE_BY_YEAR = {
    2012: "Before", 2013: "Before", 2014: "Before", 2015: "Before", 2016: "Before",
    2017: "During", 2018: "After",
}
PHASE_ORDER = ["Before", "During", "After"]


def phase_map(years, pre_scandal_end: int, scandal_end: int) -> dict[int, str]:
    """Before/During/After split from adjustable boundary years. Display only."""
    out = {}
    for y in years:
        out[y] = "Before" if y <= pre_scandal_end else ("During" if y <= scandal_end else "After")
    return out

IS_YEARS = [2018, 2017, 2016, 2015, 2014, 2013]
BS_YEARS = [2018, 2017, 2016, 2015, 2014, 2013, 2012]
CF_YEARS = [2018, 2017, 2016, 2015, 2014, 2013]


def _clean(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("\xa0", " ").strip()
    if s in {"", "-", "–", "—", "n/a", "N/A"}:
        return None
    s = s.replace(" ", "").replace(",", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return None


def _label(v):
    if v is None:
        return None
    return re.sub(r"\s+", " ", str(v).replace("\xa0", " ")).strip()


IS_MAP = {
    "Revenue": "Revenue",
    "Cost of Sales": "Cost of sales",
    "Gross Profit": "Gross profit",
    "Operating Profit/(Loss)": "Operating profit",
    "Finance Costs": "Finance costs",
    "Profit/(Loss) Before Taxation": "Profit before taxation",
    "Taxation": "Taxation",
    "Profit/(Loss) from Continuing Operations": "Profit from continuing operations",
    "Profit/(Loss) for the Year": "Profit for the year",
    "Total Comprehensive Income": "Total comprehensive income",
}

BS_MAP = {
    "Goodwill": "Goodwill",
    "Intangible assets": "Intangible assets",
    "Property, plant and equipment": "Property, plant and equipment",
    "Inventories": "Inventories",
    "Cash and cash equivalents": "Cash and cash equivalents",
    "Total assets": "Total assets (as captured)",
    "Total equity": "Total equity",
    "Total equity and liabilities": "Total equity and liabilities",
    "Accumulated losses": "Accumulated losses",
}

BS_SUBTOTALS = [
    "Total non-current assets",
    "Total current assets (excl. held for sale)",
    "Total current assets (incl. held for sale)",
    "Total non-current liabilities",
    "Total current liabilities (excl. held for sale)",
    "Total current liabilities (incl. held for sale)",
    "Total liabilities",
]

CF_MAP = {
    "Cash generated from operations": "Cash generated from operations",
    "Net cash inflow from operating activities": "Operating cash flow",
    "Net cash outflow from investing activities": "Investing cash flow",
    "Net cash inflow from financing activities": "Financing cash flow",
    "Additions to property, plant and equipment and investment property": "Capital expenditure",
    "CASH AND CASH EQUIVALENTS AT END OF YEAR": "Closing cash",
}


def parse_workbook(path: str | Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, data_only=True)
    rec: list[dict] = []

    ws = wb["Income Statement"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        std = IS_MAP.get(_label(row[0]))
        if std is None:
            continue
        for i, yr in enumerate(IS_YEARS, start=1):
            v = _clean(row[i])
            if v is not None:
                rec.append(dict(statement="income", line_item=std, fy=yr, value=v))

    ws = wb["Financial Position"]
    sub_i, ibl_i = 0, 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        lab = _label(row[0])
        std = None
        if lab is None:
            if any(_clean(c) is not None for c in row[1:8]) and sub_i < len(BS_SUBTOTALS):
                std, sub_i = BS_SUBTOTALS[sub_i], sub_i + 1
        elif lab == "Interest-bearing loans and borrowings":
            std = ["Interest-bearing borrowings (non-current)",
                   "Interest-bearing borrowings (current)"][min(ibl_i, 1)]
            ibl_i += 1
        elif lab.startswith("Assets and disposal groups"):
            std = "Assets held for sale"
        else:
            std = BS_MAP.get(lab)
        if std is None:
            continue
        for i, yr in enumerate(BS_YEARS, start=1):
            v = _clean(row[i])
            if v is not None:
                rec.append(dict(statement="position", line_item=std, fy=yr, value=v))

    ws = wb["Cash Flow"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        std = CF_MAP.get(_label(row[0]))
        if std is None:
            continue
        for i, yr in enumerate(CF_YEARS, start=1):
            v = _clean(row[i])
            if v is not None:
                rec.append(dict(statement="cashflow", line_item=std, fy=yr, value=v))

    df = pd.DataFrame.from_records(rec)
    df["currency"] = CURRENCY
    df["phase"] = df["fy"].map(PHASE_BY_YEAR)
    return df


def parse_events(path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["RedFlags_Timeline"]
    rows = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r[0]:
            continue
        rows.append(dict(period=_label(r[0]), event=_label(r[1]),
                         category=_label(r[2]), indicator=_label(r[3])))
    ev = pd.DataFrame(rows)
    if ev.empty:
        return ev

    def anchor(p):
        y = [int(x) for x in re.findall(r"(20\d{2})", p or "")]
        return max(y) if y else None

    ev["fy_anchor"] = ev["period"].map(anchor)
    ev["phase"] = ev["fy_anchor"].map(lambda y: PHASE_BY_YEAR.get(y, "After") if y else "After")
    return ev


def parse_prices(path) -> pd.DataFrame:
    """Share price anchor points. Not a series — a handful of sourced observations."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "SharePrice_Data" not in wb.sheetnames:
        return pd.DataFrame()
    rows = []
    for r in wb["SharePrice_Data"].iter_rows(min_row=3, values_only=True):
        if not r[0]:
            continue
        price = _clean(r[2])
        rows.append(dict(
            date=pd.to_datetime(str(r[0]), errors="coerce"),
            event=_label(r[1]), price=price, price_text=_label(r[2]),
            exact=price is not None, market_cap=_clean(r[3]),
            source=_label(r[4]), phase=_label(r[5]) or "Before",
        ))
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    return df.dropna(subset=["date"]).sort_values("date").reset_index(drop=True)


def parse_distress(path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Distress_Models" not in wb.sheetnames:
        return pd.DataFrame()
    years, z, zone = None, None, None
    for r in wb["Distress_Models"].iter_rows(values_only=True):
        lab = _label(r[0]) or ""
        if lab == "Year" and years is None:
            years = [int(v) for v in r[1:7] if isinstance(v, (int, float))]
        elif lab.startswith("Z''-Score"):
            z = [_clean(v) for v in r[1:7]]
        elif lab.startswith("Zone"):
            zone = [_label(v) for v in r[1:7]]
    if not years or z is None:
        return pd.DataFrame()
    n = len(years)
    return pd.DataFrame(dict(fy=years, z_score=z[:n],
                             zone=(zone or [None] * n)[:n])).set_index("fy", drop=False)


def parse_beneish(path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Distress_Models" not in wb.sheetnames:
        return pd.DataFrame()
    wanted = {"DSRI (Days-Sales-Receivables Index)": "DSRI",
              "GMI (Gross Margin Index, prior/current)": "GMI",
              "SGI (Sales Growth Index)": "SGI",
              "LVGI (Leverage Index)": "LVGI",
              "AQI (Asset Quality Index)": "AQI"}
    years, data = None, {}
    for r in wb["Distress_Models"].iter_rows(values_only=True):
        lab = _label(r[0]) or ""
        if lab == "Year":
            years = [int(v) for v in r[1:7] if isinstance(v, (int, float))]
        if lab in wanted:
            data[wanted[lab]] = [_clean(v) for v in r[1:7]]
    if not years or not data:
        return pd.DataFrame()
    out = pd.DataFrame({k: v[:len(years)] for k, v in data.items()}, index=years)
    out.index.name = "fy"
    return out


# ----------------------------------------------------------------------------
# Derived measures
# ----------------------------------------------------------------------------

def _pivot(df):
    return df.pivot_table(index="fy", columns="line_item", values="value", aggfunc="first")


def _g(p, col, fy):
    if col not in p.columns or fy not in p.index:
        return None
    v = p.at[fy, col]
    return None if pd.isna(v) else float(v)


def _div(a, b):
    if a is None or b is None or b == 0:
        return None
    return a / b


def build_ratios(statements: pd.DataFrame) -> pd.DataFrame:
    """Bases, applied consistently and stated in the report: profit for the year;
    total assets as non-current plus current including held for sale; current
    items including held for sale; debt as interest-bearing borrowings, current
    plus non-current.
    """
    p = _pivot(statements)
    rows = []
    for fy in sorted(set(statements["fy"])):
        rev, gp = _g(p, "Revenue", fy), _g(p, "Gross profit", fy)
        op, profit = _g(p, "Operating profit", fy), _g(p, "Profit for the year", fy)
        nca = _g(p, "Total non-current assets", fy)
        ca = _g(p, "Total current assets (incl. held for sale)", fy)
        cl = _g(p, "Total current liabilities (incl. held for sale)", fy)
        inv, eq = _g(p, "Inventories", fy), _g(p, "Total equity", fy)
        tl, gw = _g(p, "Total liabilities", fy), _g(p, "Goodwill", fy)
        intang = _g(p, "Intangible assets", fy)
        ocf, capex = _g(p, "Operating cash flow", fy), _g(p, "Capital expenditure", fy)

        ta = None if (nca is None or ca is None) else nca + ca
        d_nc = _g(p, "Interest-bearing borrowings (non-current)", fy) or 0.0
        d_c = _g(p, "Interest-bearing borrowings (current)", fy) or 0.0
        debt = (d_nc + d_c) if (d_nc or d_c) else None
        soft = (gw + intang) / ta if (gw is not None and intang is not None and ta) else None
        wc = None if (ca is None or cl is None) else ca - cl

        rows.append(dict(
            fy=fy, currency=CURRENCY, phase=PHASE_BY_YEAR.get(fy),
            current_ratio=_div(ca, cl),
            quick_ratio=_div(None if (ca is None or inv is None) else ca - inv, cl),
            debt_to_equity=_div(debt, eq), debt_ratio=_div(tl, ta),
            gross_margin=_div(gp, rev), operating_margin=_div(op, rev),
            net_margin=_div(profit, rev),
            return_on_assets=_div(profit, ta), return_on_equity=_div(profit, eq),
            asset_turnover=_div(rev, ta),
            goodwill_pct_assets=_div(gw, ta), soft_assets_pct=soft,
            accruals_ratio=_div(None if (profit is None or ocf is None) else profit - ocf, ta),
            cash_conversion=_div(ocf, op) if (op or 0) > 0 else None,
            working_capital=wc, wc_to_assets=_div(wc, ta),
            total_assets=ta, total_equity=eq, total_liabilities=tl,
            revenue=rev, profit=profit, operating_profit=op,
            operating_cash_flow=ocf, capex=capex, debt=debt,
        ))
    return pd.DataFrame(rows).set_index("fy", drop=False)


def run_checks(statements: pd.DataFrame, prices: pd.DataFrame) -> pd.DataFrame:
    p = _pivot(statements)
    f = []
    for fy in sorted(set(statements[statements.statement == "position"]["fy"])):
        nca = _g(p, "Total non-current assets", fy)
        ca = _g(p, "Total current assets (incl. held for sale)", fy)
        stated = _g(p, "Total assets (as captured)", fy)
        tel = _g(p, "Total equity and liabilities", fy)
        derived = None if (nca is None or ca is None) else nca + ca

        if stated is not None and tel is not None and abs(stated - tel) > 0.5:
            f.append(dict(fy=fy, severity="Critical", check="Balance sheet does not balance",
                          detail=(f"Total assets {stated:,.0f} against total equity and "
                                  f"liabilities {tel:,.0f}. Difference {stated - tel:+,.0f}. "
                                  "The workbook discloses this on its own balance check row.")))
        if derived is not None and stated is not None and abs(derived - stated) > 0.5:
            f.append(dict(fy=fy, severity="High",
                          check="Captured total assets inconsistent with its components",
                          detail=(f"Components sum to {derived:,.0f}; captured total reads "
                                  f"{stated:,.0f}.")))

    rev = {fy: _g(p, "Revenue", fy) for fy in sorted(set(statements["fy"]))}
    yrs = [y for y in sorted(rev) if rev.get(y)]
    for a, b, c in zip(yrs, yrs[1:], yrs[2:]):
        if rev[b] < 0.6 * rev[a] and rev[c] > 1.6 * rev[b]:
            f.append(dict(fy=b, severity="High", check="Revenue discontinuity",
                          detail=(f"Revenue falls from {rev[a]:,.0f} to {rev[b]:,.0f} and recovers "
                                  f"to {rev[c]:,.0f}. Confirm the period length and reporting "
                                  f"entity for FY{b}.")))

    cf = sorted(set(statements[statements.statement == "cashflow"]["fy"]))
    missing = [y for y in yrs if y not in cf]
    if missing:
        f.append(dict(fy=max(missing), severity="High", check="Cash flow series incomplete",
                      detail=(f"Cash flow captured for {cf}. Missing {missing}, so the accruals "
                              "test and cash conversion cannot be computed for those years.")))

    if prices is None or prices.empty:
        f.append(dict(fy=None, severity="Blocking", check="Share price not captured",
                      detail="Required for sub-question 4.4."))
    else:
        exact = int(prices["exact"].sum())
        f.append(dict(fy=None, severity="Blocking",
                      check="Share price captured as anchor points, not a series",
                      detail=(f"{len(prices)} sourced observations, {exact} carrying an exact "
                              "price. Enough to chart the collapse for 4.2, but a machine "
                              "learning model needs a continuous daily or weekly series. "
                              "Interpolating between anchors would fabricate the very data the "
                              "models are then evaluated on.")))

    for miss, sub in [("Dividends per share", "4.5"), ("Peer company multiples", "4.5")]:
        f.append(dict(fy=None, severity="Blocking", check=f"{miss} not captured",
                      detail=f"Required for sub-question {sub}."))

    return pd.DataFrame(f)


SCORECARD = [
    ("Current ratio", "current_ratio", lambda v: v >= 1.5, lambda v: v >= 1.0,
     "Below 1.0, short-term obligations exceed short-term assets."),
    ("Quick ratio", "quick_ratio", lambda v: v >= 1.0, lambda v: v >= 0.7,
     "Strips out inventory."),
    ("Debt to equity", "debt_to_equity", lambda v: v <= 1.0, lambda v: v <= 2.0,
     "Gearing against the buffer available to absorb write-downs."),
    ("Debt ratio", "debt_ratio", lambda v: v <= 0.5, lambda v: v <= 0.65,
     "Share of assets funded by liabilities."),
    ("Operating margin", "operating_margin", lambda v: v >= 0.08, lambda v: v >= 0.03,
     "Operating performance before financing and tax."),
    ("Return on equity", "return_on_equity", lambda v: 0.05 <= v <= 0.30, lambda v: v > 0,
     "An unusually high return can reflect a shrinking equity base."),
    ("Goodwill share of assets", "goodwill_pct_assets", lambda v: v <= 0.15, lambda v: v <= 0.25,
     "Acquisition-driven growth and write-down exposure."),
    ("Soft assets share", "soft_assets_pct", lambda v: v <= 0.30, lambda v: v <= 0.45,
     "Goodwill plus intangibles — the easiest assets to overstate."),
    ("Cash conversion", "cash_conversion", lambda v: v >= 0.8, lambda v: v >= 0.4,
     "Operating cash flow against operating profit."),
    ("Working capital / assets", "wc_to_assets", lambda v: v >= 0.05, lambda v: v >= 0.0,
     "Negative working capital is a going-concern signal."),
]


def build_scorecard(ratios: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for label, col, ok, warn, note in SCORECARD:
        for fy in ratios.index:
            v = ratios.at[fy, col] if col in ratios.columns else None
            if v is None or pd.isna(v):
                status = "No data"
            elif ok(v):
                status = "Healthy"
            elif warn(v):
                status = "Warning"
            else:
                status = "Critical"
            rows.append(dict(indicator=label, fy=int(fy),
                             value=None if (v is None or pd.isna(v)) else float(v),
                             status=status, note=note, phase=PHASE_BY_YEAR.get(fy)))
    return pd.DataFrame(rows)


def load_all(path: str | Path):
    statements = parse_workbook(path)
    ratios = build_ratios(statements)
    events = parse_events(path)
    prices = parse_prices(path)
    checks = run_checks(statements, prices)
    scorecard = build_scorecard(ratios)
    distress = parse_distress(path)
    beneish = parse_beneish(path)
    return statements, ratios, checks, events, scorecard, prices, distress, beneish


if __name__ == "__main__":
    import sys
    src = sys.argv[1] if len(sys.argv) > 1 else "Steinhoff_-Dashboard_Data-Version_2.xlsx"
    s, r, c, e, sc, pr, d, b = load_all(src)
    pd.set_option("display.width", 220)
    print(f"lines={len(s)}  years={sorted(set(s.fy))}")
    print(r[["current_ratio", "debt_to_equity", "net_margin", "return_on_equity",
             "goodwill_pct_assets", "cash_conversion", "accruals_ratio"]].round(3))
    print("\nZ'':"); print(d)
    print("\nBeneish:"); print(b.round(3) if not b.empty else "none")
    print(f"\nprices={len(pr)} events={len(e)}")
    for _, row in c.iterrows():
        print(f"  [{row.severity:8}] {row.fy if row.fy else '   -'}  {row.check}")
